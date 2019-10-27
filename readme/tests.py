#coding:utf-8
from django.test import TestCase

from lib.info_map import *
# info_map = InfoMap()
# print (info_map)
# info_map.add('aa',{ 'message':'1234' })
# info_map.add('aa',{ 'message':'1234' })
# print (info_map.get_map())
#
# print (info_map.pop('aa'))
# print (info_map.get_map())
# print (info_map.pop('aa'))
# print (info_map.get_map())
import datetime



import openpyxl
wb = openpyxl.load_workbook(r'C:\lab\git\uni/kejiting_uni/readme/kej_data1.xlsx')
ws = wb.active
all_list = []
for row in ws.rows:
    sub_list = []
    for col in row:
        sub_list.append(col.value)
    all_list.append(sub_list)
# print (len(all_list))

# print (type(datetime.datetime(2020, 9, 1, 0, 0)))

# print (datetime.datetime(2020, 9, 1, 0, 0).strftime("%Y-%m-%d")  )
# print (all_list)
# print (all_list.decode("utf-8",'ignore') )



# newList = []
# for sub in all_list:
#     arr = []
#     for node in sub:
#         if( type(node) == type(datetime.datetime(2020, 9, 1, 0, 0)) ):
#             node = node.strftime("%Y-%m-%d")
#         if(node is None):
#             node = ""
#             # print (datetime.datetime(2020, 9, 1, 0, 0).strftime("%Y-%m-%d")  )
#         arr.append(node)
#     newList.append(arr)
# print (newList)


newList = []
for sub in all_list:
    arr = []
    arr.append(sub[0]) #id   0
    arr.append(sub[2]) #项目名称  1
    arr.append(sub[3]) #公司名称  2
    arr.append(sub[4]) #负责人  3
    arr.append(sub[15]) #完成率  4
    arr.append(sub[25]) #项目申请时间  5
    newList.append(arr)
print (newList)





        # arr.append(node)
        # print (node)
        # print( type(node) )
        # if type(node) == "str":
        #     # n =  node
        #     print( node )
        # else:
        #
        #     try:
        #         # pass
        #         print(node)
        #     except:
        #         print(sub[0] )
        #         raise



        # print(col.value)



# 从工作薄中获取一个表单(sheet)对象
sheets = wb.sheetnames
# print(sheets, type(sheets))


# print('{}行 {}列'.format(ws.max_row, ws.max_column))
# 创建一个表单
# mySheet = wb.create_sheet('mySheet')
# print(wb.sheetnames)
#
# # 获取指定的表单
# sheet3 = wb.get_sheet_by_name('Sheet3')
# sheet4 = wb['mySheet']
# for sheet in wb:
#     print(sheet.title)




# import requests
# url = 'https://www.xiaohongshu.com/discovery/item/5d25f0f8000000002701243f'
# headers = {'content-type': 'application/json'}
# r = requests.get(url, headers=headers )
# print(r)
# f=open('red.html',"wb")
# f.write(r.content)
# f.close()
# Create your tests here.
# INFO_QUEUE = {}
# INFO_QUEUE['aa'] = [ {'message':'me'}, {'message':'u1'},]
# INFO_QUEUE['bb'] = [ {'message':'me123'}, {'message':'1123'},]
#
# def get_value(key):
#     value = INFO_QUEUE.pop(key,None)
#     if value is None :
#         return { 'message':u'none'}
#     else:
#         return {'value':value}
# key = 'aa'
# v = get_value(key)
# print (v)
# print (INFO_QUEUE)